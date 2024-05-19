from shutil import copy2
import customtkinter as ctk
from tkinter import filedialog, messagebox, Label
import openpyxl
import cv2  # pip install opencv-python
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import img2pdf
from PIL import Image, ImageDraw, ImageFont, ImageTk
import numpy as np
from pdf2image import convert_from_path
from pathsAndMail import *


def dodajSzablon():
    global templatePath
    templatePath = filedialog.askopenfilename()  # dodac pulpit

    print(templatePath)

    try:
        # Convert PDF to JPG
        pdf_path = templatePath
        output_images = convert_from_path(pdf_path, poppler_path=popplerPath)
        output_images_paths = []
        for i, image in enumerate(output_images):
            image_path = f'page{i}.png'
            image.save(image_path, 'PNG')
            output_images_paths.append(image_path)

        if templatePath != programPath:
            copy2(templatePath, templatesFolderPath)
            return templatePath
    except:
        messagebox.showerror('Błąd', 'Nie wczytano pliku!')


def dodajPlikXLSX():
    try:
        global xlsxPath
        xlsxPath = filedialog.askopenfilename()
        print(xlsxPath)
        return xlsxPath
    except:
        messagebox.showerror('Błąd', 'Coś poszło nie tak!')


def generowanieCertyfikatow():
    try:
        global path, wb_obj, sheet_obj, script_dir, output_folder
        path = xlsxPath
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active
        script_dir = os.path.dirname(os.path.abspath(__file__))
        output_folder = os.path.join(script_dir, 'Wygenerowane zaswiadczenia')
        os.makedirs(output_folder, exist_ok=True)
        # Iterate through rows in the Excel file
        for a in range(2, sheet_obj.max_row + 1):
            imie_obj = sheet_obj.cell(row=a, column=1)
            nazwisko_obj = sheet_obj.cell(row=a, column=2)
            nr_obj = sheet_obj.cell(row=a, column=10)
            kurs_obj = sheet_obj.cell(row=a, column=3)
            data_obj = sheet_obj.cell(row=a, column=4)
            miejsce_obj = sheet_obj.cell(row=a, column=5)
            daty_obj = sheet_obj.cell(row=a, column=8)

            # DODAWANIE TEKSTU DO PDF-A
            img_path = f'page0.png'
            img = cv2.imread(img_path)

            # Kordy imie i nazwisko
            top_left_x = 113
            top_left_y = 675
            org = (top_left_x, top_left_y)

            # Kordy nr zaswiadczenia
            top_left_x1 = 572
            top_left_y1 = 504
            org1 = (top_left_x1, top_left_y1)

            # Kordy kursu
            top_left_x2 = 112
            top_left_y2 = 114
            org2 = (top_left_x2, top_left_y2)

            # Kordy daty urodzenia
            top_left_x3 = 110
            top_left_y3 = 768
            org3 = (top_left_x3, top_left_y3)

            # Kordy dat
            top_left_x4 = 112
            top_left_y4 = 1165
            org4 = (top_left_x4, top_left_y4)

            # Uczestniczyl
            top_left_x5 = 113
            top_left_y5 = 933
            org5 = (top_left_x5, top_left_y5)

            # Write text
            font_path = programPath + '\\fontStyle\\BAHNSCHRIFT.TTF'
            font_size = 75  # Replace with the desired font size
            font_size1 = 36
            font_size2 = 54

            bahnschrift_font = ImageFont.truetype(font_path, font_size)
            bahnschrift_font1 = ImageFont.truetype(font_path, font_size1)
            bahnschrift_font2 = ImageFont.truetype(font_path, font_size2)
            color = (84, 84, 84)

            text = str(imie_obj.value) + " " + str(nazwisko_obj.value)
            text1 = "Nr " + str(nr_obj.value)
            text2 = str(kurs_obj.value)
            text3 = str(data_obj.value) + " w " + str(miejsce_obj.value)
            text4 = str(daty_obj.value)
            text5 = "uczeniczyl"

            # Create a PIL image from the OpenCV image
            pil_image = Image.fromarray(cv2.cvtColor(img, cv2.COLOR_BGR2RGB))

            # Create a PIL draw object
            draw = ImageDraw.Draw(pil_image)

            # Write text with "Bahnschrift" font using PIL
            draw.text(org, text.upper(), font=bahnschrift_font, fill=color)
            draw.text(org1, text1, font=bahnschrift_font2, fill=color)
            draw.text(org2, text2, font=bahnschrift_font1, fill=color)
            draw.text(org3, text3, font=bahnschrift_font1, fill=color)
            draw.text(org4, text4, font=bahnschrift_font1, fill=color)
            draw.text(org5, text5, font=bahnschrift_font1, fill=color)

            # Convert the PIL image back to a NumPy array
            img_with_text = cv2.cvtColor(np.array(pil_image), cv2.COLOR_RGB2BGR)

            # Save the resulting image
            filename = f'{imie_obj.value} {nazwisko_obj.value} - certyfikat.png'
            filepath = os.path.join(output_folder, filename)
            cv2.imwrite(filepath, img_with_text)

            # PNG TO PDF
            image_to_convert = filepath
            pdf_filename = f'{imie_obj.value} {nazwisko_obj.value} - certyfikat.pdf'
            pdf_filepath = os.path.join(output_folder, pdf_filename)

            # Convert the redacted image to PDF
            with open(pdf_filepath, "wb") as file:
                file.write(img2pdf.convert(image_to_convert))

            # Output
            print(f"Successfully made PDF file {pdf_filename}")
            for a in range(2, sheet_obj.max_row + 1):
                png_filename = f'{sheet_obj.cell(row=a, column=1).value} {sheet_obj.cell(row=a, column=2).value} - certyfikat.png'
                png_filepath = os.path.join(output_folder, png_filename)
                if os.path.exists(png_filepath):
                    os.remove(png_filepath)
    except:
        messagebox.showerror('Błąd', 'Coś poszło nie tak. Spróbuj ponownie !')


def sendMail():
    global text
    # MAIL
    body = '''Hello,
    This is the body of the email
    sincerely yours
    G.G.
    '''
    sender = senderEmail
    password = senderPassword
    receiver = receiverEmail
    for a in range(2, sheet_obj.max_row + 1):
        #receiver = sheet_obj.cell(row=a, column=9)
        # Set up the MIME
        message = MIMEMultipart()
        message['From'] = sender
        message['To'] = receiver
        message['Subject'] = 'This email has an attachment, a PDF file'
        message.attach(MIMEText(body, 'plain'))
        filename = f'{sheet_obj.cell(row=a, column=1).value} {sheet_obj.cell(row=a, column=2).value} - certyfikat.pdf'
        filepath = os.path.join(output_folder, filename)
        # Open the file in binary mode
        binary_pdf = open(filepath, 'rb')
        payload = MIMEBase('application', 'octate-stream', Name=filename)
        payload.set_payload(binary_pdf.read())
        # Encode the binary into base64
        encoders.encode_base64(payload)
        # Add header with PDF name
        payload.add_header('Content-Decomposition', 'attachment', filename=filename)
        message.attach(payload)
        # Use Gmail with port
        session = smtplib.SMTP('smtp.gmail.com', 587)
        # Enable security
        session.starttls()
        # Login with email and password
        session.login(sender, password)
        text = message.as_string()
        session.sendmail(sender, receiver, text)
        session.quit()
        print('Mail Sent')


if __name__ == '__main__':
    ctk.deactivate_automatic_dpi_awareness()  ### skalowanie
    window = ctk.CTk()  # traktowane jako główne okno
    # window.iconbitmap("myIcon.ico")
    window.title('Generator Dyplomów')
    window.geometry('1200x700')

    # configure grid layout (4x4)
    window.grid_columnconfigure((1, 2), weight=1)
    window.grid_rowconfigure((0, 1, 2, 3, 4, 5), weight=2)

    # create sidebar frame with widgets
    window.sidebar_frame = ctk.CTkFrame(window, width=240, corner_radius=0, height=1200)
    window.sidebar_frame.grid(row=0, column=0, rowspan=4, sticky="nsew")
    window.sidebar_frame.grid_rowconfigure(4, weight=1)
    window.logo_label = ctk.CTkLabel(window.sidebar_frame, text="Opcje",
                                     font=ctk.CTkFont(size=40, weight="bold"))  # stary font 20
    window.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))
    window.sidebar_button_1 = ctk.CTkButton(window.sidebar_frame, text="Dodaj Szablon", command=dodajSzablon,
                                            font=ctk.CTkFont(size=20, weight="normal"), width=190, height=25)
    window.sidebar_button_1.grid(row=1, column=0, padx=25, pady=10)
    window.sidebar_button_2 = ctk.CTkButton(window.sidebar_frame, text="Wybierz plik xlsx", command=dodajPlikXLSX,
                                            font=ctk.CTkFont(size=20, weight="normal"), width=190, height=25)
    window.sidebar_button_2.grid(row=2, column=0, padx=20, pady=10)
    window.sidebar_button_3 = ctk.CTkButton(window.sidebar_frame, text="Generuj", command=generowanieCertyfikatow,
                                            font=ctk.CTkFont(size=20, weight="normal"), width=190, height=25)
    window.sidebar_button_3.grid(row=3, column=0, padx=20, pady=10)
    window.sidebar_button_3 = ctk.CTkButton(window.sidebar_frame, text="Wyślij na maila", command=sendMail,
                                            font=ctk.CTkFont(size=20, weight="normal"), width=190, height=25)
    window.sidebar_button_3.grid(row=5, column=0, padx=20, pady=10)
    ###############################
    # load images with light and dark mode image
    image_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "test_images")
    window.large_test_image = ctk.CTkImage(Image.open(os.path.join(image_path, pdfLookup)), size=(200, 200))

    # Create an object of tkinter ImageTk
    img = ImageTk.PhotoImage(Image.open(pdfLookup))

    # Create a Label Widget to display the text or Image
    label = Label(window, image=img)
    label.grid(row=1, column=1, padx=20, pady=10)
    window.mainloop()
